# -*- coding: utf-8 -*-
"""
Reporte por contrato — CLINICA ANESHI WAYAA IPS SAS (NIT 900653282)
===================================================================
Genera un Excel con 3 hojas para los contratos indicados:

  1. "Contratos"      -> contratos del prestador (NIT 900653282)
  2. "Facturas"       -> facturas radicadas de los contratos: nro factura,
                          nro de radicacion, valor, fecha factura, fecha de
                          radicacion y fecha de prestacion del servicio.
  3. "Autorizaciones" -> radicados + listado de autorizaciones (consulta
                          oficial del area), filtrado por estos contratos.

Optimizacion: las autorizaciones de los 3 contratos se identifican PRIMERO
(CTE aut_sel) y la subconsulta de detalle (CUP/medicamentos/insumos) se
restringe solo a esas autorizaciones -> de minutos a segundos, sin cambiar
los resultados.

Requisitos:
    pip install psycopg2-binary pandas openpyxl

Uso:
    python reporte_contratos_aneshi.py
"""

import os
import sys
import warnings
import pandas as pd
import psycopg2
from openpyxl.utils import get_column_letter

# La lectura con conexion DBAPI2 (psycopg2) funciona; solo silenciamos el aviso.
warnings.filterwarnings("ignore", message="pandas only supports SQLAlchemy")

# ── CONFIGURACION DE CONEXION ────────────────────────────────────────────────
DB_CONFIG = {
    "host":     "129.213.136.60",
    "port":     6432,
    "dbname":   "base_sie_dusakawi",
    "user":     "postgres",
    "password": "qazwsx12A.",
    "sslmode":  "disable",
}

# ── PARAMETROS DEL REPORTE ───────────────────────────────────────────────────
NIT_PRESTADOR = "900653282"
CONTRATOS = [
    "EV-44078-2026-50",
    "EV-44078-2026-48-PQ",
    "EV-44078-2026-12-MD",
]
# Ruta de salida (absoluta). La carpeta se crea automaticamente si no existe.
ARCHIVO_SALIDA = r"C:\reportes\Reporte_ANESHI_WAYAA_contratos.xlsx"

# ── Mapeo numero_contrato -> consecutivo_contrato (filtro rapido por entero) ──
SQL_CONSECUTIVOS = """
SELECT consecutivo_contrato
FROM administrativo.ct_ips_contrato
WHERE numero_contrato = ANY(%(contratos)s);
"""

# ── 1) CONTRATOS DEL PRESTADOR ───────────────────────────────────────────────
SQL_CONTRATOS = """
SELECT
    c.numero_contrato,
    c.consecutivo_contrato,
    i.nit,
    i.razon_social                         AS prestador,
    i.codigo_habilitacion,
    c.fecha_inicio,
    c.fecha_terminacion,
    c.valor_contrato,
    c.monto_ejecutado,
    CASE c.estado
         WHEN 1 THEN 'En elaboracion'
         WHEN 2 THEN 'Legalizado'
         WHEN 3 THEN 'Activo'
         ELSE 'Estado ' || c.estado::text
    END                                    AS estado_contrato,
    CASE WHEN c.sw_activo = 1 THEN 'SI' ELSE 'NO' END AS activo
FROM administrativo.ct_ips_contrato c
JOIN administrativo.ct_ips i ON i.ips = c.ips
WHERE i.nit = %(nit)s
ORDER BY c.fecha_inicio DESC, c.numero_contrato;
"""

# ── 2) FACTURAS RADICADAS DE LOS CONTRATOS ───────────────────────────────────
SQL_FACTURAS = """
SELECT
    f.numero_contrato,
    f.numero_factura,
    f.radicacion                           AS numero_radicacion,
    f.fecha                                AS fecha_factura,
    f.fecha_radicacion_inicio              AS fecha_radicacion,
    f.fecha_servicio_inicio                AS fecha_prestacion_inicio,
    f.fecha_servicio_fin                   AS fecha_prestacion_fin,
    f.valor_factura                        AS valor_servicios,
    f.valor_copago,
    f.saldo_factura                        AS saldo,
    f.consecutivo_rips_af,
    CASE f.estado_factura
         WHEN 1 THEN 'Activa/Radicada'
         WHEN 2 THEN 'Anulada'
         ELSE 'Estado ' || f.estado_factura::text
    END                                    AS estado_factura,
    i.razon_social                         AS prestador,
    i.nit
FROM administrativo.sc_factura_encabezado f
LEFT JOIN administrativo.ct_ips i ON i.ips = f.consecutivo_ips
WHERE f.numero_contrato = ANY(%(contratos)s)
ORDER BY f.numero_contrato, f.fecha, f.numero_factura;
"""

# ── 3) RADICADOS + LISTADO DE AUTORIZACIONES (consulta oficial, optimizada) ──
#  Cambios vs. la original:
#   - CTE aut_sel: autorizaciones de los contratos (por consecutivo_contrato).
#   - La subconsulta T (CUP/medicamento/insumo) se restringe a aut_sel.
#   - WHERE final filtra por consecutivo_contrato (entero, indexado) en vez de
#     por el rango de fechas / numero_autorizacion.
#  Los resultados son identicos a la consulta original aplicada a estos contratos.
SQL_AUTORIZACIONES = """
with aut_sel as (
    select consecutivo_autorizacion
    from administrativo.ss_autorizacion
    where consecutivo_contrato = ANY(%(consecutivos)s)
      and consecutivo_autorizacion is not null
      and sw_orden_servicio = 0
)
select distinct
sol.numero_solicitud,
aut.consecutivo_interno as numero_autorizacion,
to_char(sol.fecha_solicitud_medico, 'dd/mm/yyyy hh24:mi') fecha_de_solicitud_de_orden_medica,
to_char(aut.fecha_grabado, 'dd/mm/yyyy hh24:mi') as fecha_autorizacion,
to_char(aut.fecha_real_prestacion_servicio, 'dd/mm/yyyy hh24:mi') as fecha_prestacion_del_servicio,
to_char(aut.fecha_real_autorizacion, 'dd/mm/yyyy hh24:mi') as fecha_activacion_autorizacion,
to_char(sol.fecha_solicitud, 'dd/mm/yyyy hh24:mi') as fecha_solicitud_de_autorizacion,
extract(days from (aut.fecha_grabado - sol.fecha_solicitud)) as Oportunidad_en_la_generacion_de_autorizaciones,
extract(days from (cast(aut.fecha_real_prestacion_servicio as date) - aut.fecha_grabado)) as Oportunidad_en_la_prestacion_del_servicio,
aut.tipo_identificacion_afiliado, aut.numero_identificacion_afiliado,
concat(aut.primer_nombre_afiliado, ' ', aut.segundo_nombre_afiliado, ' ', aut.primer_apellido_afiliado, ' ', aut.segundo_apellido_afiliado) as nombre_completo_afiliado,
aut.primer_nombre_afiliado, aut.segundo_nombre_afiliado, aut.primer_apellido_afiliado, aut.segundo_apellido_afiliado,
aut.fecha_nacimiento_afiliado,
round(extract(days from (current_timestamp - cast(aut.fecha_nacimiento_afiliado as timestamp)  ))/365)  edad_afiliado,
af.fecha_afiliacion_sgsss fecha_afiliacion_adres,
case when aut.estado_afiliado=1 then 'Activo' when  aut.estado_afiliado=2 then 'Retirado' when  aut.estado_afiliado=3 then 'Fallecido' end estado_de_afiliacion,
af.telefono_1 telefono_afiliado,
af.direccion direccion_de_residencia,
etnia.descripcion etnia,
asentamiento.descripcion Asentamiento_comunidad_o_Rancheria,
case when af.sexo = 1 then 'M'
when af.sexo = 2 then 'F'
else 'O'
end as sexo,
case when af.zona_afiliacion = 1 then 'R'
when af.zona_afiliacion = 2 then 'U'
end as zona,
depA.municipio as codigo_departamento_afiliado,
depA.descripcion as departamento_afiliado,
munA.municipio as codigo_municipio_afiliado,
munA.descripcion as municipio_afiliado,
aut.nivel_sisben_afiliado,
pobA.descripcion as grupo_poblacional,
case when aut.tipo_regimen_afiliado = 1 then 'Contributivo'
when aut.tipo_regimen_afiliado = 99 then 'Subsidiado'
end as regimen_afiliado,
case when aut.tipo_regimen_afiliado = 1 then 'EPSIC1'
when aut.tipo_regimen_afiliado = 99 then 'EPSI01'
end as codigo_epsi,
sol.nit_ips as nit_ips_que_solicita_servicio,
munS.descripcion as municipio_que_solicita_servicio,
sol.razon_social_ips as razon_social_ips_solicita,
esp.descripcion as especialidad_que_solicita_el_servicio,
aut.nit_prestador as nit_ips_destino_prestadora_servicio,
aut.razon_social_prestador as ips_destino_prestadora_del_servicio,
espAut.descripcion ESPECILIADAD_QUE_PRESTA_EL_SERVICIO,
munP.descripcion as muncipio_destino,
cont.numero_contrato numero_contrato_entidad_destino,
cont.fecha_inicio as fecha_inicio,
cont.fecha_terminacion as fecha_terminacion,
nivel.homologacion as nivel_complejidad,
case when aut.sw_portabilidad = 0 then 'NO'
else 'SI' end as portabilidad,
T.codigo_interno as Codigo_Cups_Cum_Autorizado,
T.descripcion_codigo_propio as Descripcion_del_servicio_autorizado,
T.grupo_servicio Codigo_grupo_servicio,
T.descripcion_servicio Descripcion_grupo_de_servicio,
T.cantidad as Cantidad_de_servicios_autorizados,
ssm.dias medicamentos_dias_de_tratamiento,
T.valor_tarifario as Valor_unitario_de_servicios_autorizados,
T.valor_servicio as Valor_total_de_los_servicios_autorizados,
cie.simbolo as cie_10,
cie.descripcion as Descripcion_del_diagnostico_principal,
cie2.simbolo as cie_10_Secundario,
cie2.descripcion as Descripcion_del_diagnostico_secundario,
'' Diagnostico_programa_de_riesgo,
case when sol.ubicacion_paciente = 1 then 'Ambulatorio'
when sol.ubicacion_paciente = 2 then 'Urgencias'
when sol.ubicacion_paciente = 3 then 'Hospitalizacion'
when sol.ubicacion_paciente = 4 then 'Domiciliario'
end as ambito_del_servicio,
case when usuariosol.usuario is not null then upper(concat (usuariosol.usuario,'_', usuariosol.nombre, ' ', usuariosol.apellido)) else 'IPS Solicita' end as Nombre_de_quien_solicita_el_servicio,
upper(concat (usuario.usuario,'_', usuario.nombre, ' ', usuario.apellido)) as Nombre_de_quien_autoriza_el_servicio,
case when aut.fecha_anula is not null then 'Anulado'
when aut.fecha_real_prestacion_servicio is not null then 'Activa con Prestacion'
when aut.fecha_real_autorizacion is not null then 'Activa sin Prestacion'
when aut.pin is not null then 'Sin Activar'
else 'Solicitud en Tramite'
end as Estado_de_autorizacion,
to_char(aut.fecha_anula, 'dd/mm/yyyy hh24:mi') as Fecha_de_anulacion_de_la_autorizacion,
concat (usuario_anula.usuario,'_', usuario_anula.nombre, ' ', usuario_anula.apellido) as Usuario_quien_anula_la_autorizacion,
aut.observacion_anula,
aut.fecha_autorizacion_reserva as Fecha_de_reserva_tecnica,
T.observacion_autorizacion observacion_autorizacion_de_servicios,
case when aut.consecutivo_rips is not null then 'SI'
else 'NO' end as estado_radicacion,
case when aut.tipo_cobro in (1,2) then aut.valor_copago else 0 end valor_copago,
to_char(sol.fecha_grabado, 'dd/mm/yyyy hh24:mi') Fecha_de_Radicacion_de_la_solicitud_ante_la_eps,
case when aut.tipo_cobro not in (1,2) then aut.valor_copago else 0 end valor_cuotamoderadora,
iss.codigo_habilitacion as habilita_solicita,
iss.codigo_prestador as prestador_solicita,
ia.codigo_habilitacion as habilita_autoriza,
ia.codigo_prestador as prestador_autoriza,
ci.razon_social as IpsPrimaria
from administrativo.ss_solicitud sol
left join administrativo.ss_autorizacion aut
on sol.consecutivo_solicitud = aut.consecutivo_solicitud
left join administrativo.ct_ips iss on sol.ips_solicitante = iss.ips
left join administrativo.ct_ips ia on aut.consecutivo_ips = ia.ips
left join administrativo.af_afiliado af on aut.afiliado = af.afiliado
left join administrativo.ct_ips ci on af.ips = ci.ips
left join administrativo.tb_indigena_etnia etnia on af.indigena_etnia =etnia.consecutivo
left join administrativo.tb_indigena_asentamiento asentamiento on af.indigena_acentamiento  =asentamiento.consecutivo
left join administrativo.tb_municipio munA
on aut.municipio_afiliado = munA.municipio
left join administrativo.tb_municipio depA
on munA.departamento = depA.municipio
left join administrativo.tb_grupo_poblacional pobA
on af.grupo_poblacional = pobA.grupo_poblacional
left join administrativo.tb_municipio munS
on sol.municipio_solicitante = munS.municipio
left join administrativo.tb_municipio munP
on aut.municipio_prestador = munP.municipio
left join administrativo.ct_ips_contrato cont
on aut.consecutivo_contrato = cont.consecutivo_contrato
left join administrativo.tb_nivel nivel
on nivel.consecutivo_nivel = aut.consecutivo_nivel
left join administrativo.usuario usuario
on aut.usuario_grabado = usuario.usuario
left join administrativo.usuario usuariosol
on sol.usuario_grabado = usuariosol.usuario
left join administrativo.usuario usuario_anula
on aut.usuario_anula = usuario_anula.usuario
left join
(
	select cupA.consecutivo_autorizacion, 0 consecutivo, cup.codigo_interno, cupA.descripcion_codigo_propio, cupA.cantidad, cupA.valor_servicio, cupA.fecha_cancelacion, cupA.fecha_prestacion_servicio, cupA.observacion_autorizacion,
	tec.codigo_interno as grupo_servicio, tec.descripcion as descripcion_servicio, cupA.valor_tarifario
	from administrativo.ss_autorizacion_cup cupA
	left join administrativo.tb_cup cup
	on cupA.cup = cup.cup
	left join administrativo.tb_cup_nota_tecnica nota
	on nota.consecutivo_cup = cup.cup
	left join administrativo.tb_concepto_nota_tecnica tec
	on tec.consecutivo_concepto = nota.consecutivo_concepto
	where cupA.consecutivo_autorizacion in (select consecutivo_autorizacion from aut_sel)
	union all
	select medA.consecutivo_autorizacion, medA.medicamento consecutivo, med.codigo_interno, medA.descripcion_codigo_propio, medA.cantidad, medA.valor_servicio, medA.fecha_cancelacion, medA.fecha_prestacion_servicio, medA.observacion_autorizacion,
	tec.codigo_interno as grupo_servicio, tec.descripcion as descripcion_servicio, medA.valor_tarifario
	from administrativo.ss_autorizacion_medicamento medA
	left join administrativo.tb_medicamento med
	on medA.medicamento = med.medicamento
	left join administrativo.tb_medicamento_nota_tecnica nota
	on nota.consecutivo_medicamento = med.medicamento
	left join administrativo.tb_concepto_nota_tecnica tec
	on tec.consecutivo_concepto = nota.consecutivo_concepto
	where medA.consecutivo_autorizacion in (select consecutivo_autorizacion from aut_sel)
	union all
	select insA.consecutivo_autorizacion, 0 consecutivo, ins.codigo_interno, insA.descripcion_codigo_propio, insA.cantidad, insA.valor_servicio, insA.fecha_cancelacion, insA.fecha_prestacion_servicio, insA.observacion_autorizacion,
	tec.codigo_interno as grupo_servicio, tec.descripcion as descripcion_servicio, insA.valor_tarifario
	from administrativo.ss_autorizacion_insumo insA
	left join administrativo.tb_insumo ins
	on insA.insumo = ins.insumo
	left join administrativo.tb_insumo_nota_tecnica nota
	on nota.consecutivo_insumo = ins.insumo
	left join administrativo.tb_concepto_nota_tecnica tec
	on tec.consecutivo_concepto = nota.consecutivo_concepto
	where insA.consecutivo_autorizacion in (select consecutivo_autorizacion from aut_sel)
) T
on T.consecutivo_autorizacion = aut.consecutivo_autorizacion
left join administrativo.ss_solicitud_medicamento ssm on sol.consecutivo_solicitud =ssm.consecutivo_solicitud and ssm.consecutivo_autorizacion = aut.consecutivo_autorizacion and ssm.consecutivo_medicamento = T.consecutivo
left join administrativo.tb_cie10 cie
on cie.cie_10 = sol.diagnostico_principal
left join administrativo.tb_cie10 cie2
on cie2.cie_10 = sol.diagnostico_relacionado_1
left join administrativo.tb_especialidad esp
on sol.consecutivo_especialidad = esp.consecutivo_especialidad
left join administrativo.tb_especialidad espAut
on espAut.consecutivo_especialidad = aut.consecutivo_especialidad
where aut.consecutivo_contrato = ANY(%(consecutivos)s)
and aut.consecutivo_autorizacion is not null
and aut.sw_orden_servicio = 0
"""


def ejecutar(conn, sql, params, etiqueta):
    """Ejecuta una consulta y devuelve un DataFrame; informa progreso."""
    print(f"  -> Consultando {etiqueta} ...", flush=True)
    df = pd.read_sql_query(sql, conn, params=params)
    print(f"     {len(df):,} filas.", flush=True)
    return df


def auto_formato(writer, hoja, df, columnas_moneda=()):
    """Congela encabezado, agrega autofiltro, ajusta ancho y formatea moneda."""
    ws = writer.sheets[hoja]
    ws.freeze_panes = "A2"
    if len(df):
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"
    for idx, col in enumerate(df.columns, start=1):
        letra = get_column_letter(idx)
        muestra = df[col].astype(str).head(500)
        ancho = max([len(str(col))] + [len(v) for v in muestra]) + 2
        ws.column_dimensions[letra].width = min(max(ancho, 10), 60)
        if col in columnas_moneda:
            for fila in range(2, len(df) + 2):
                ws.cell(row=fila, column=idx).number_format = '#,##0'


def main():
    print("Conectando a la base de datos...", flush=True)
    try:
        conn = psycopg2.connect(**DB_CONFIG, connect_timeout=15)
    except Exception as e:
        print(f"ERROR de conexion: {e}", file=sys.stderr)
        sys.exit(1)

    try:
        # consecutivos de los contratos (filtro rapido para autorizaciones)
        df_consec = pd.read_sql_query(SQL_CONSECUTIVOS, conn, params={"contratos": CONTRATOS})
        consecutivos = [int(x) for x in df_consec["consecutivo_contrato"].tolist()]
        print(f"  Contratos objetivo (consecutivos): {consecutivos}", flush=True)

        df_contratos = ejecutar(conn, SQL_CONTRATOS, {"nit": NIT_PRESTADOR}, "contratos del prestador")
        df_facturas  = ejecutar(conn, SQL_FACTURAS,  {"contratos": CONTRATOS}, "facturas radicadas")
        df_autoriz   = ejecutar(conn, SQL_AUTORIZACIONES, {"consecutivos": consecutivos}, "autorizaciones (radicados)")
    finally:
        conn.close()

    # Asegura que la carpeta de destino exista
    carpeta = os.path.dirname(ARCHIVO_SALIDA)
    if carpeta:
        os.makedirs(carpeta, exist_ok=True)

    print(f"\nEscribiendo Excel: {ARCHIVO_SALIDA} ...", flush=True)
    with pd.ExcelWriter(ARCHIVO_SALIDA, engine="openpyxl") as writer:
        df_contratos.to_excel(writer, sheet_name="Contratos", index=False)
        df_facturas.to_excel(writer,  sheet_name="Facturas",  index=False)
        df_autoriz.to_excel(writer,   sheet_name="Autorizaciones", index=False)

        auto_formato(writer, "Contratos", df_contratos,
                     columnas_moneda=("valor_contrato", "monto_ejecutado"))
        auto_formato(writer, "Facturas", df_facturas,
                     columnas_moneda=("valor_servicios", "valor_copago", "saldo"))
        auto_formato(writer, "Autorizaciones", df_autoriz,
                     columnas_moneda=("valor_unitario_de_servicios_autorizados",
                                      "valor_total_de_los_servicios_autorizados",
                                      "valor_copago", "valor_cuotamoderadora"))

    print("\nResumen:")
    print(f"  Contratos del prestador : {len(df_contratos):,}")
    print(f"  Facturas                : {len(df_facturas):,}")
    print(f"  Filas de autorizaciones : {len(df_autoriz):,}")
    print(f"\nListo. Archivo generado en:\n  {os.path.abspath(ARCHIVO_SALIDA)}")


if __name__ == "__main__":
    main()
