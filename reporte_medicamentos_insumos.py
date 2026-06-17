# -*- coding: utf-8 -*-
"""
Reporte de Medicamentos e Insumos entregados por departamento y programa
========================================================================
Reproduce el cuadro de "Medicamentos e Insumos" (PBS por enfermedad + Insumos)
por departamento (Cesar, Magdalena, La Guajira), para el mes indicado.

DEFINICION USADA (documentada, ajustable):
  - Fuente "Entregados": medicamentos de autorizaciones CON prestacion en el
    mes (ss_autorizacion_medicamento.fecha_prestacion_servicio). Misma base de
    la consulta oficial de autorizaciones.
  - Departamento: el del afiliado (aut.municipio_afiliado -> tb_municipio).
  - Cantidad: SUMA de unidades (cantidad).
  - Clasificacion por enfermedad: CIE-10 del diagnostico principal de la
    solicitud (sol.diagnostico_principal -> tb_cie10.simbolo), por rangos:
        HTA          : I10-I15
        Diabetes     : E10-E14
        Cancer       : C00-C97  (+ D00-D09 in situ)
        VIH/SIDA     : B20-B24, Z21
        ERC          : N17-N19  (renal cronica N18)
        Autoinmunes  : M05, M06, M30-M36
        Hemofilia    : D65-D69
  - Insumos: ss_autorizacion_insumo con prestacion en el mes.

NOTA: es una definicion de mejor esfuerzo; si el area maneja otra regla
(por programa de riesgo o por dispensacion RIPS), se ajusta el SQL.

Requisitos:  pip install psycopg2-binary pandas openpyxl
Uso:         python reporte_medicamentos_insumos.py
"""

import os
import sys
import warnings
import pandas as pd
import psycopg2
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", message="pandas only supports SQLAlchemy")

# ── CONEXION ─────────────────────────────────────────────────────────────────
DB_CONFIG = {
    "host":     "129.213.136.60",
    "port":     6432,
    "dbname":   "base_sie_dusakawi",
    "user":     "postgres",
    "password": "qazwsx12A.",
    "sslmode":  "disable",
}

# ── PARAMETROS ───────────────────────────────────────────────────────────────
ANIO = 2026
MES  = 4                     # 4 = abril
DEPARTAMENTOS = ['%CESAR%', '%MAGDALENA%', '%GUAJIRA%']
ARCHIVO_SALIDA = r"C:\reportes\Medicamentos_Insumos_{anio}_{mes:02d}.xlsx"

MESES_ES = {1:'enero',2:'febrero',3:'marzo',4:'abril',5:'mayo',6:'junio',
            7:'julio',8:'agosto',9:'septiembre',10:'octubre',11:'noviembre',12:'diciembre'}

# ── SQL: medicamentos por enfermedad (cantidades) ────────────────────────────
SQL_MEDICAMENTOS = """
WITH base_med AS (
  SELECT depA.descripcion AS departamento,
         medA.cantidad::numeric AS cantidad,
         upper(left(coalesce(cie.simbolo,''),3)) AS cie3,
         upper(left(coalesce(cie.simbolo,''),1)) AS cie1
  FROM administrativo.ss_autorizacion_medicamento medA
  JOIN administrativo.ss_autorizacion aut ON aut.consecutivo_autorizacion = medA.consecutivo_autorizacion
  JOIN administrativo.ss_solicitud   sol ON sol.consecutivo_solicitud     = aut.consecutivo_solicitud
  LEFT JOIN administrativo.tb_cie10  cie ON cie.cie_10 = sol.diagnostico_principal
  JOIN administrativo.tb_municipio  munA ON aut.municipio_afiliado = munA.municipio
  JOIN administrativo.tb_municipio  depA ON munA.departamento      = depA.municipio
  WHERE medA.fecha_prestacion_servicio >= %(ini)s
    AND medA.fecha_prestacion_servicio <  %(fin)s
    AND depA.descripcion ILIKE ANY(%(deptos)s)
)
SELECT departamento,
 SUM(cantidad)                                                                            AS "Medicamentos PBS Entregados",
 COALESCE(SUM(cantidad) FILTER (WHERE cie3 IN ('I10','I11','I12','I13','I14','I15')),0)    AS "Medicamentos PBS HTA",
 COALESCE(SUM(cantidad) FILTER (WHERE cie3 IN ('E10','E11','E12','E13','E14')),0)          AS "Medicamentos PBS Diabetes",
 COALESCE(SUM(cantidad) FILTER (WHERE cie1='C' OR cie3 LIKE 'D0%%'),0)                     AS "Medicamentos PBS Cancer",
 COALESCE(SUM(cantidad) FILTER (WHERE cie3 IN ('B20','B21','B22','B23','B24','Z21')),0)    AS "Medicamentos PBS VIH/SIDA",
 COALESCE(SUM(cantidad) FILTER (WHERE cie3 IN ('N17','N18','N19')),0)                      AS "Medicamentos PBS Enfermedad Renal Cronica",
 COALESCE(SUM(cantidad) FILTER (WHERE cie3 IN ('M05','M06','M30','M31','M32','M33','M34','M35','M36')),0) AS "Medicamentos PBS Autoinmunes y Artritis",
 COALESCE(SUM(cantidad) FILTER (WHERE cie3 IN ('D65','D66','D67','D68','D69')),0)          AS "Medicamentos PBS Hemofilia y Coagulopatias"
FROM base_med
GROUP BY departamento;
"""

# ── SQL: insumos / dispositivos (cantidades) ─────────────────────────────────
SQL_INSUMOS = """
SELECT depA.descripcion AS departamento,
       SUM(insA.cantidad::numeric) AS "Dispositivos Medicos e Insumos"
FROM administrativo.ss_autorizacion_insumo insA
JOIN administrativo.ss_autorizacion aut ON aut.consecutivo_autorizacion = insA.consecutivo_autorizacion
JOIN administrativo.tb_municipio munA ON aut.municipio_afiliado = munA.municipio
JOIN administrativo.tb_municipio depA ON munA.departamento      = depA.municipio
WHERE insA.fecha_prestacion_servicio >= %(ini)s
  AND insA.fecha_prestacion_servicio <  %(fin)s
  AND depA.descripcion ILIKE ANY(%(deptos)s)
GROUP BY depA.descripcion;
"""

# Orden y tipo de cada categoria, igual al cuadro original
CATEGORIAS_PBS = [
    "Medicamentos PBS Entregados",
    "Medicamentos PBS HTA",
    "Medicamentos PBS Diabetes",
    "Medicamentos PBS Cancer",
    "Medicamentos PBS VIH/SIDA",
    "Medicamentos PBS Enfermedad Renal Cronica",
    "Medicamentos PBS Autoinmunes y Artritis",
    "Medicamentos PBS Hemofilia y Coagulopatias",
]
CATEGORIA_INSUMO = "Dispositivos Medicos e Insumos"


def main():
    ini = f"{ANIO}-{MES:02d}-01"
    fin = f"{ANIO + (MES // 12)}-{(MES % 12) + 1:02d}-01"   # primer dia del mes siguiente
    periodo = f"{MESES_ES[MES]}-{ANIO}"
    params = {"ini": ini, "fin": fin, "deptos": DEPARTAMENTOS}

    print(f"Periodo: {periodo}  ({ini} a {fin})", flush=True)
    print("Conectando...", flush=True)
    try:
        conn = psycopg2.connect(**DB_CONFIG, connect_timeout=15)
    except Exception as e:
        print(f"ERROR de conexion: {e}", file=sys.stderr)
        sys.exit(1)

    try:
        df_med = pd.read_sql_query(SQL_MEDICAMENTOS, conn, params=params)
        df_ins = pd.read_sql_query(SQL_INSUMOS, conn, params=params)
    finally:
        conn.close()

    # Unir y pasar a formato largo (una fila por Departamento + Categoria)
    df = df_med.merge(df_ins, on="departamento", how="outer").fillna(0)

    filas = []
    for _, r in df.sort_values("departamento").iterrows():
        depto = r["departamento"]
        for cat in CATEGORIAS_PBS:
            filas.append([depto, periodo, "Medicamentos e Insumos", "PBS", cat, int(round(r.get(cat, 0)))])
        filas.append([depto, periodo, "Medicamentos e Insumos", "Insumos",
                      CATEGORIA_INSUMO, int(round(r.get(CATEGORIA_INSUMO, 0)))])

    salida = pd.DataFrame(filas, columns=[
        "Departamento", "Periodo", "Grupo", "Tipo", "Categoria", "Cantidad"])

    archivo = ARCHIVO_SALIDA.format(anio=ANIO, mes=MES)
    carpeta = os.path.dirname(archivo)
    if carpeta:
        os.makedirs(carpeta, exist_ok=True)

    print(f"Escribiendo: {archivo}", flush=True)
    with pd.ExcelWriter(archivo, engine="openpyxl") as writer:
        salida.to_excel(writer, sheet_name="Medicamentos_Insumos", index=False)
        ws = writer.sheets["Medicamentos_Insumos"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(salida.columns))}{len(salida)+1}"
        anchos = [16, 14, 24, 10, 44, 12]
        for i, w in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for fila in range(2, len(salida) + 2):           # formato miles a Cantidad
            ws.cell(row=fila, column=6).number_format = '#,##0'

    print(f"\nListo. Archivo generado en:\n  {os.path.abspath(archivo)}")
    print(f"\nResumen {periodo}:")
    for _, r in salida[salida['Categoria'] == 'Medicamentos PBS Entregados'].iterrows():
        print(f"  {r['Departamento']:<12} Entregados: {r['Cantidad']:,}")


if __name__ == "__main__":
    main()
